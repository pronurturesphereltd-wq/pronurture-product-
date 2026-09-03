"""Parsing a staff list out of an uploaded CSV or Excel file.

Kept free of Django and network calls so it can be tested directly against
bytes. The spec named openpyxl for both formats, but openpyxl reads .xlsx
only — CSV goes through the standard library instead.
"""

import csv
import io

from openpyxl import load_workbook

CSV_EXTENSIONS = {".csv"}
EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
SUPPORTED_EXTENSIONS = CSV_EXTENSIONS | EXCEL_EXTENSIONS

REQUIRED_COLUMNS = ["full_name", "email", "license_number", "license_body"]
OPTIONAL_COLUMNS = ["phone"]

# Spreadsheets in the wild rarely use our exact field names.
# Keys must be written as they look *after* normalise_header lowercases the
# text and turns hyphens into spaces — so "e mail", never "e-mail".
COLUMN_ALIASES = {
    "full name": "full_name",
    "name": "full_name",
    "fullname": "full_name",
    "email address": "email",
    "e mail": "email",
    "email addr": "email",
    "licence_number": "license_number",
    "licence number": "license_number",
    "license number": "license_number",
    "registration number": "license_number",
    "licence_body": "license_body",
    "licence body": "license_body",
    "license body": "license_body",
    "regulator": "license_body",
    "phone number": "phone",
    "mobile": "phone",
    "telephone": "phone",
}


class ImportFileError(Exception):
    """The uploaded file cannot be parsed at all."""


def extension_of(filename):
    name = (filename or "").lower()
    dot = name.rfind(".")
    return name[dot:] if dot != -1 else ""


def normalise_header(value):
    key = str(value or "").strip().lower().replace("-", " ")
    key = " ".join(key.split())
    if key in COLUMN_ALIASES:
        return COLUMN_ALIASES[key]
    return key.replace(" ", "_")


def parse_staff_file(content, filename):
    """Return (rows, headers). Each row is a dict of column -> string value.

    Raises ImportFileError if the file is unreadable or missing a required
    column. Individual bad rows are not rejected here — that is row-level
    validation, and the import reports those per row rather than failing whole.
    """
    extension = extension_of(filename)
    if extension in CSV_EXTENSIONS:
        table = _read_csv(content)
    elif extension in EXCEL_EXTENSIONS:
        table = _read_excel(content)
    else:
        raise ImportFileError(
            f"Unsupported file type '{extension or filename}'. "
            f"Use one of: {', '.join(sorted(SUPPORTED_EXTENSIONS))}."
        )

    if not table:
        raise ImportFileError("The file is empty.")

    headers = [normalise_header(cell) for cell in table[0]]
    missing = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing:
        raise ImportFileError(
            "Missing required column(s): "
            + ", ".join(missing)
            + f". Found: {', '.join(h for h in headers if h) or '(none)'}."
        )

    rows = []
    for values in table[1:]:
        row = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            value = values[index] if index < len(values) else ""
            row[header] = "" if value is None else str(value).strip()
        # Skip the blank trailing rows spreadsheets love to produce.
        if any(row.get(column) for column in REQUIRED_COLUMNS):
            rows.append(row)
    return rows, headers


def _read_csv(content):
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ImportFileError("Could not decode the CSV file as text.")

    try:
        return [row for row in csv.reader(io.StringIO(text))]
    except csv.Error as exc:
        raise ImportFileError(f"Could not parse the CSV file: {exc}") from exc


def _read_excel(content):
    try:
        workbook = load_workbook(
            io.BytesIO(content), read_only=True, data_only=True
        )
    except Exception as exc:
        raise ImportFileError(f"Could not open the Excel file: {exc}") from exc

    try:
        sheet = workbook.worksheets[0]
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()
